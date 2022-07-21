#!python3
# excel_to_csv.py - search the current working directory for any Excel files,
#                   and save them as CSV files. If the Excel file has multiple
#                   sheets, each sheet gets its own CSV file


import csv
import os
import openpyxl
from openpyxl.utils import get_column_letter


def get_csv_from_excel() -> None:
    """
    Loop through the current working directory and create an Excel workbook object for any
    Excel files. Loop through the workbook object to create a sheet object. Create csv_file_object
    and pass it to the csv_writer_object. Loop through the rows and columns of the sheet object.
    Append the cells to row_data. Write row_data to a CSV file then close the csv_file_object. A
    new file object will be created on the next iteration.
    """
    for excel_file in os.listdir("."):
        if not excel_file.endswith(".xlsx"):
            continue
        wb = openpyxl.load_workbook(excel_file)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            csv_file_name = excel_file.split(".")[0] + f"_{sheet_name}.csv"
            csv_file_object = open(csv_file_name, "w", newline="")
            csv_writer_object = csv.writer(csv_file_object)
            for row_num in range(1, sheet.max_row + 1):
                row_data = []  # append each cell to this list
                for col_num in range(1, sheet.max_column + 1):
                    row_data.append(
                        sheet[get_column_letter(col_num) + str(row_num)].value
                    )
                csv_writer_object.writerow(row_data)
            csv_file_object.close()


if __name__ == "__main__":
    get_csv_from_excel()
