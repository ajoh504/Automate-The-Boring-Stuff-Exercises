#!python3
# texttosheet.py - read multiple text files and store their contents in an Excel
#                  file. The first text file goes in column A, and each line in
#                  goes into the rows of column A.

import sys
import openpyxl
from openpyxl.utils import get_column_letter


class TextToSheet:
    def __init__(self, text_files: list):
        """
        get_lines() returns a list of lines from a .txt file

        get_list_of_lines() calls get_lines() and returns a list of nested
        lists. Each nested list contains the list of lines from get_lines()
        """
        self.get_lines = lambda file: open(file, "r").readlines()
        self.get_list_of_lines = lambda: [
            self.get_lines(file) for file in self.text_files
        ]
        self.text_files = text_files  # sys.orig_argv minus the first two args
        self.wb = openpyxl.Workbook()  # Excel workbook object
        self.sheet = self.wb.active  # Excel sheet object

    def write_to_sheet(self) -> None:
        for i, list in enumerate(self.get_list_of_lines()):
            for j, line in enumerate(list):
                self.sheet[get_column_letter(i + 1) + str(j + 1)].value = line
        self.wb.save("text_added.xlsx")


if __name__ == "__main__":
    text_to_sheet = TextToSheet(sys.orig_argv[2:])  # skip the first two sys args
    text_to_sheet.write_to_sheet()
