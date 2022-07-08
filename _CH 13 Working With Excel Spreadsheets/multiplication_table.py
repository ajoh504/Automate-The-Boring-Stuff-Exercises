#!python3
# multiplication_table.py -- take a number (num) from the command line and create a
# multiplication table in Excel with the contents num x num. Make the first row and
# column bold.

import sys
import openpyxl
from openpyxl.styles import Font

def create_multiplication_table() -> None:
    '''
    Create two for loops to represent row and column numbers. Add 2 to each number to
    account for skipping the 1x1 cell, and to account for the first row and column (i.e. the 
    grid numbers). For cells with row and column numbers > 1, add multiplication table elements.

    '''
    num = int(sys.argv[1])
    wb = openpyxl.Workbook() # create Workbook object
    sheet = wb.active # create Workbook sheet object
    bold = Font(bold=True)

    for row_num in range(1, num + 2):
        for column_number in range(1, num + 2):
            # make first column with bold font
            if row_num == 1 and column_number > 1:
                sheet.cell(row=row_num, column=column_number).value = column_number - 1
                sheet.cell(row=row_num, column=column_number).font = bold
            # make first row with bold font
            elif column_number == 1 and row_num > 1:
                sheet.cell(row=row_num, column=column_number).value = row_num - 1
                sheet.cell(row=row_num, column=column_number).font = bold
            # create multiplication table
            elif column_number > 1 and row_num > 1:
                sheet.cell(row=row_num, column=column_number).value = (row_num - 1) * (column_number - 1)
    wb.save('mult_table_' + str(num) + 'x' + str(num) + '.xlsx')

if __name__ == '__main__':
    create_multiplication_table()
