#!python3
# blank_row_inserter.py - take two numbers and an Excel filename as command line
#                         arguments. The first number denotes a row number. The second
#                         number denotes how many blank rows to insert starting at the
#                         given row number.

import sys
import openpyxl
from openpyxl.utils import get_column_letter


class BlankRowInserter:
    def __init__(self, start: int, qty: int, excel_file: str):
        """
        :param start: blank row starting point (sys.argv[1])
        :param qty: quantity of blank rows to add (sys.argv[2])
        :param excel_file: file to add blank rows to (sys.argv[3]
        """
        self.start = start
        self.qty = qty
        self.excel_file = excel_file
        self.wb = openpyxl.load_workbook(self.excel_file)  # Excel workbook object
        self.sheet = self.wb.active  # Excel sheet object

    def get_new_cells(self) -> dict:
        """
        Append Excel cell coordinates/values to dictionary keys/values. At
        the blank row starting pount, add None as values to create blank lines.
        Continue adding None for as many times as designated by self.qty.
        Append all the rest of the Excel cell values to the new_cells dictionary.

        :return new_cells: dictionary with blank rows added as None value type
        """
        new_cells = {}
        for i in range(1, self.sheet.max_row + self.qty + 1):
            for j in range(1, self.sheet.max_column + 1):
                # add values before blank rows
                if i < self.start:
                    new_cells[get_column_letter(j) + str(i)] = self.sheet[
                        get_column_letter(j) + str(i)
                    ].value
                # add blank rows
                elif i == self.start or i < (self.start + self.qty):
                    new_cells[get_column_letter(j) + str(i)] = None
                # add values after blank rows
                elif i >= self.start + self.qty:
                    new_cells[get_column_letter(j) + str(i)] = self.sheet[
                        get_column_letter(j) + str(i - self.qty)
                    ].value
        return new_cells

    def blank_row_inserter(self, new_cells: dict) -> None:
        """
        Loop through row and column numbers. Append new_cells values to Excel
        cells. Save as new file.
        """
        for i in range(1, self.sheet.max_row + self.qty + 1):
            for j in range(1, self.sheet.max_column + 1):
                self.sheet[get_column_letter(j) + str(i)].value = new_cells[
                    get_column_letter(j) + str(i)
                ]
        self.wb.save(self.excel_file[0:-5] + "_edited" + ".xlsx")


if __name__ == "__main__":
    class_object = BlankRowInserter(int(sys.argv[1]), int(sys.argv[2]), sys.argv[3])
    class_object.blank_row_inserter(class_object.get_new_cells())
