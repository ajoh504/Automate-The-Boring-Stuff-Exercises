#!python3
# sheettotext.py - move all columns from an Excel sheet into their own
#                  individual .txt files. Supply the Excel file as a
#                  command line argument.

import sys
import openpyxl
from openpyxl.utils import get_column_letter


def write_sheet_to_file() -> None:
    """
    sys.argv[1] must contain an Excel file name.

    :exception TypeError: None values in Excel cells will throw a TypeError
    in the open() function. Except the TypeError and write '' to file instead
    of None.
    """
    wb = openpyxl.load_workbook(sys.argv[1])
    sheet = wb.active
    for column in range(1, sheet.max_column + 1):
        for row in range(1, sheet.max_row + 1):
            try:
                with open("column" + get_column_letter(column) + ".txt", "a") as file:
                    file.write(sheet[get_column_letter(column) + str(row)].value)
            except TypeError:
                with open("column" + get_column_letter(column) + ".txt", "a") as file:
                    file.write("")


if __name__ == "__main__":
    write_sheet_to_file()
